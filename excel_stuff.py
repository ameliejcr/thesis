import openpyxl
import numpy as np

wb_phon = openpyxl.load_workbook('eng_phonetic_frame.xlsx', True)
wb_sem = openpyxl.load_workbook('semantic_frame.xlsx', True)

ws_phon_train = wb_phon['Sheet1']
ws_sem_train = wb_sem['Sheet1']
ws_phon_test = wb_phon['Sheet2']
ws_sem_test = wb_sem['Sheet2']

phon_lib_train = []
sem_lib_train = []
phon_lib_test = []
sem_lib_test = []
input_list_phon_train = []
input_list_sem_train = []
input_list_phon_test = []
input_list_sem_test = []


for row in ws_phon_train.iter_rows(min_row=5, min_col=2, max_col=145, max_row=642, values_only=True):
    phon_lib_train.append(row)

for row in ws_sem_train.iter_rows(min_row=2, min_col=2, max_col=401, max_row=639, values_only=True):
    sem_lib_train.append(row)

for row in ws_phon_test.iter_rows(min_row=1, min_col=2, max_col=145, max_row=30, values_only=True):
    phon_lib_test.append(row)

for row in ws_sem_test.iter_rows(min_row=1, min_col=2, max_col=401, max_row=30, values_only=True):
    sem_lib_test.append(row)

for row in ws_phon_test.iter_rows(min_row=1, min_col=2, max_col=11, max_row=10, values_only=True):
    input_list_phon_train.append(row)

for row in ws_sem_test.iter_rows(min_row=1, min_col=2, max_col=11, max_row=10, values_only=True):
    input_list_sem_train.append(row)

for row in ws_phon_test.iter_rows(min_row=11, min_col=2, max_col=11, max_row=21, values_only=True):
    input_list_phon_test.append(row)

for row in ws_sem_test.iter_rows(min_row=11, min_col=2, max_col=11, max_row=21, values_only=True):
    input_list_sem_test.append(row)

row_3 = ws_phon_train.iter_rows(min_row=3, min_col=2, max_row=3, max_col=145)

phon_dict = {'place': np.array([0.1, 0.2, 0.3, 0.4, 0.5, 0.6, 0.7, 0.8, 0.9, 1.0]),
             'manner': np.array([0.0, 0.1, 0.3, 0.5, 0.8, 1.0]),
             'voiced': np.array([0.5, 1.0]),
             'lateral': np.array([0.0, 0.5]),
             'open': np.array([0.0, 0.4, 0.5, 0.6, 0.8, 1.0]),
             'front': np.array([0.0, 0.1, 0.5, 0.9, 1.0]),
             'long': np.array([0.5, 1.0]),
             'rounded': np.array([0.0, 0.5])}

def get_closest(output):
    new_output = []
    for i in output:
        if output.index(i) == 0 | 4 | 16 | 20| 24| 28| 40| 44| 48| 52 | 64| 68|72|76|88|92|96|100|112|116|120|124|136|140:
            dict_key = phon_dict['place']
            idx = (np.abs(dict_key - i)).argmin()
            new_output.append(dict_key[idx])
        elif output.index(i) == 1|5|17|21|25|29|41|45|49|53|65|69|73|77|89|93|97|101|113|117|121|125|137|141:
            dict_key = phon_dict['manner']
            idx = (np.abs(dict_key - i)).argmin()
            new_output.append(dict_key[idx])
        elif output.index(i) == 2|6|18|22|26|30|42|46|50|54|66|70|74|78|90|94|98|102|114|118|122|126|138|142:
            dict_key = phon_dict['voiced']
            idx = (np.abs(dict_key - i)).argmin()
            new_output.append(dict_key[idx])
        elif output.index(i) == 3|7|19|23|27|31|43|47|51|55|67|71|75|79|91|95|99|103|115|119|123|127|139|143:
            dict_key = phon_dict['lateral']
            idx = (np.abs(dict_key - i)).argmin()
            new_output.append(dict_key[idx])
        elif output.index(i) == 8|12|32|36|56|60|80|84|96|104|108|128|132:
            dict_key = phon_dict['open']
            idx = (np.abs(dict_key - i)).argmin()
            new_output.append(dict_key[idx])
        elif output.index(i) == 9|13|33|37|57|61|81|85|97|105|109|129|133:
            dict_key = phon_dict['front']
            idx = (np.abs(dict_key - i)).argmin()
            new_output.append(dict_key[idx])
        elif output.index(i) == 10|14|34|38|58|62|82|86|98|106|110|130|134:
            dict_key = phon_dict['long']
            idx = (np.abs(dict_key - i)).argmin()
            new_output.append(dict_key[idx])
        elif output.index(i) == 11|15|35|39|59|63|83|87|99|107|111|131|135:
            dict_key = phon_dict['rounded']
            idx = (np.abs(dict_key - i)).argmin()
            new_output.append(dict_key[idx])

